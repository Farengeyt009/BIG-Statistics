"""
Сервис экспорта задач проекта в Excel (.xlsx)
"""
from io import BytesIO
from typing import List, Dict, Any, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Переводы заголовков ─────────────────────────────────────────────────────
HEADERS_I18N: Dict[str, List[str]] = {
    'en': ['ID', 'Title', 'Description', 'Status', 'Priority',
           'Assignee', 'Start Date', 'Due Date', 'Created'],
    'zh': ['ID', '标题', '描述', '状态', '优先级',
           '负责人', '开始日期', '截止日期', '创建日期'],
    'ru': ['ID', 'Название', 'Описание', 'Статус', 'Приоритет',
           'Исполнитель', 'Дата начала', 'Дедлайн', 'Создано'],
}

# ── Переводы приоритетов ────────────────────────────────────────────────────
PRIORITY_I18N: Dict[str, Dict[str, str]] = {
    'en': {'low': 'Low', 'medium': 'Medium', 'high': 'High',
           'critical': 'Critical', 'urgent': 'Urgent'},
    'zh': {'low': '低', 'medium': '中', 'high': '高',
           'critical': '严重', 'urgent': '紧急'},
    'ru': {'low': 'Низкий', 'medium': 'Средний', 'high': 'Высокий',
           'critical': 'Критический', 'urgent': 'Срочный'},
}

# ── Переводы стандартных статусов ───────────────────────────────────────────
# Ключи — английские названия, которые система создаёт по умолчанию.
# Кастомные (не попавшие сюда) выгружаются как есть.
STATUS_I18N: Dict[str, Dict[str, str]] = {
    'en': {
        # Английские оригиналы
        'Backlog': 'Backlog', 'To Do': 'To Do', 'In Progress': 'In Progress',
        'In Review': 'In Review', 'Done': 'Done', 'Cancelled': 'Cancelled', 'New': 'New',
        # Русские варианты
        'Новый': 'New', 'Новая': 'New', 'К выполнению': 'To Do', 'В работе': 'In Progress',
        'На проверке': 'In Review', 'Выполнено': 'Done', 'Завершена': 'Done',
        'Отменено': 'Cancelled', 'Отменена': 'Cancelled',
        # Китайские варианты
        '新建': 'New', '待办': 'To Do', '进行中': 'In Progress',
        '审核中': 'In Review', '完成': 'Done', '取消': 'Cancelled',
    },
    'zh': {
        # Английские оригиналы
        'Backlog': '待规划', 'To Do': '待办', 'In Progress': '进行中',
        'In Review': '审核中', 'Done': '完成', 'Cancelled': '取消', 'New': '新建',
        # Русские варианты
        'Новый': '新建', 'Новая': '新建', 'К выполнению': '待办', 'В работе': '进行中',
        'На проверке': '审核中', 'Выполнено': '完成', 'Завершена': '完成',
        'Отменено': '取消', 'Отменена': '取消',
        # Китайские оригиналы
        '新建': '新建', '待办': '待办', '进行中': '进行中',
        '审核中': '审核中', '完成': '完成', '取消': '取消',
    },
    'ru': {
        # Английские оригиналы
        'Backlog': 'Беклог', 'To Do': 'К выполнению', 'In Progress': 'В работе',
        'In Review': 'На проверке', 'Done': 'Завершена', 'Cancelled': 'Отменена', 'New': 'Новый',
        # Русские оригиналы (возвращаем как есть)
        'Новый': 'Новый', 'Новая': 'Новая', 'К выполнению': 'К выполнению', 'В работе': 'В работе',
        'На проверке': 'На проверке', 'Выполнено': 'Выполнено', 'Завершена': 'Завершена',
        'Отменено': 'Отменено', 'Отменена': 'Отменена',
        # Китайские варианты
        '新建': 'Новый', '待办': 'К выполнению', '进行中': 'В работе',
        '审核中': 'На проверке', '完成': 'Завершена', '取消': 'Отменена',
    },
}

SHEET_TITLE: Dict[str, str] = {
    'en': 'Tasks', 'zh': '任务', 'ru': 'Задачи',
}

# Ширины стандартных колонок (без "Тип"): ID + 8 полей
COL_WIDTHS = [8, 40, 50, 18, 14, 22, 14, 14, 14]


def _fmt_date(val: Optional[str]) -> str:
    if not val:
        return ''
    return str(val)[:10]


def _build_cf_rows_for_task(task_id: int, cf_defs: List[Dict], cf_values: Dict) -> List[Dict]:
    """
    Возвращает список строк кастомных полей для одной задачи.
    cf_values: { task_id -> { row_index -> { field_id -> value } } }
    """
    task_rows = cf_values.get(task_id, {})
    if not task_rows:
        return [{f['id']: '' for f in cf_defs}]

    result = []
    for row_index in sorted(task_rows.keys()):
        row_vals = task_rows[row_index]
        result.append({f['id']: row_vals.get(f['id'], '') for f in cf_defs})
    return result


def _task_values(task: Dict, lang: str) -> List[Any]:
    priorities = PRIORITY_I18N.get(lang, PRIORITY_I18N['en'])
    statuses = STATUS_I18N.get(lang, STATUS_I18N['en'])
    raw_status = task.get('status_name') or ''
    translated_status = statuses.get(raw_status, raw_status)
    assignee = task.get('assignee_full_name') or task.get('assignee_name') or ''
    return [
        task.get('id'),
        task.get('title', ''),
        task.get('description') or '',
        translated_status,
        priorities.get(task.get('priority', ''), task.get('priority', '')),
        assignee,
        _fmt_date(task.get('start_date')),
        _fmt_date(task.get('due_date')),
        _fmt_date(task.get('created_at')),
    ]


def export_tasks_to_excel(
    tasks: List[Dict],
    cf_defs: List[Dict],
    cf_values: Dict,
    lang: str = 'ru',
) -> BytesIO:
    """
    Строит .xlsx файл и возвращает его как BytesIO.

    tasks     — список задач проекта (только корневые)
    cf_defs   — список определений кастомных полей проекта
    cf_values — { task_id: { row_index: { field_id: value } } }
    lang      — язык интерфейса: 'ru' | 'en' | 'zh'
    """
    lang = lang if lang in HEADERS_I18N else 'en'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_TITLE.get(lang, 'Tasks')

    # ── Заголовки ──────────────────────────────────────────────────────────────
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(fill_type='solid', fgColor='3B5BDB')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    cf_headers = [f['field_name'] for f in cf_defs]
    all_headers = HEADERS_I18N[lang] + cf_headers
    ws.append(all_headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[1].height = 24

    # ── Стили строк ────────────────────────────────────────────────────────────
    center = Alignment(horizontal='center', vertical='center')
    wrap_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def write_row(values: List[Any]):
        ws.append(values)
        r = ws.max_row
        for col_idx, cell in enumerate(ws[r], start=1):
            cell.alignment = center if col_idx == 1 else wrap_left
            cell.border = thin_border
        ws.row_dimensions[r].height = 30.75

    # ── Запись строк (только корневые задачи) ──────────────────────────────────
    root_tasks = [t for t in tasks if not t.get('parent_task_id')]

    for task in root_tasks:
        task_id = task['id']
        cf_row_list = _build_cf_rows_for_task(task_id, cf_defs, cf_values)
        std = _task_values(task, lang)

        write_row(std + list(cf_row_list[0].values()))

        # Дополнительные строки если CF multi-row
        for cf_row in cf_row_list[1:]:
            extra = [''] * len(HEADERS_I18N[lang]) + list(cf_row.values())
            extra[0] = task_id  # оставляем ID
            write_row(extra)

    # ── Ширина колонок ─────────────────────────────────────────────────────────
    for col_idx in range(1, ws.max_column + 1):
        width = COL_WIDTHS[col_idx - 1] if col_idx <= len(COL_WIDTHS) else 20
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
